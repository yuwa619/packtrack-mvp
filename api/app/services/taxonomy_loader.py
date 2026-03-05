from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from ..db.models import TaxonomyCode
from ..schemas.defra import TAXONOMY_SHEET_NAME


@dataclass
class TaxonomySeedResult:
    inserted: int
    sheet_name: str


def load_taxonomy_from_excel(session: Session, excel_path: Path) -> TaxonomySeedResult:
    workbook = load_workbook(excel_path, data_only=True)
    if TAXONOMY_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Sheet '{TAXONOMY_SHEET_NAME}' not found in {excel_path}")

    sheet = workbook[TAXONOMY_SHEET_NAME]

    session.execute(delete(TaxonomyCode))

    inserted = 0
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=3, values_only=True), start=2
    ):
        category, code, description = row
        if not category or not code or not description:
            continue

        session.add(
            TaxonomyCode(
                category=str(category).strip(),
                code=str(code).strip(),
                description=str(description).strip(),
                source_sheet=TAXONOMY_SHEET_NAME,
                source_row_number=row_index,
                active=True,
            )
        )
        inserted += 1

    return TaxonomySeedResult(inserted=inserted, sheet_name=TAXONOMY_SHEET_NAME)
