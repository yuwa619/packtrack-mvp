from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from api.app.schemas.defra import DEFRA_REPORT_COLUMNS, TAXONOMY_SHEET_NAME


def _resolve_workbook_path() -> Path:
    candidates = [
        Path("data/defra/UK DEFRA.xlsx"),
        Path("data/defra/UK_DEFRA.xlsx"),
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("DEFRA workbook not found")


def test_report_columns_match_workbook_header_exactly() -> None:
    workbook = load_workbook(_resolve_workbook_path(), data_only=True)
    sheet = workbook["example of a populated DEFRA EP"]
    header = [sheet.cell(1, index).value for index in range(1, 16)]

    assert header == DEFRA_REPORT_COLUMNS


def test_taxonomy_sheet_entry_count() -> None:
    workbook = load_workbook(_resolve_workbook_path(), data_only=True)
    sheet = workbook[TAXONOMY_SHEET_NAME]
    rows = [
        row
        for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True)
        if any(value is not None and str(value).strip() for value in row)
    ]

    assert len(rows) == 46
